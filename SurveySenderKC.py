import shutil
import openpyxl
import time
import sys
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
import PySimpleGUI as sg
import phonenumbers

import phonenumber_extractor
import webform_filler

#Inspiration = https://realpython.com/pysimplegui-python/ 

# Define the Excel file path and sheet name
# folder_directory = Path( __file__ ).parent.absolute() # gets current folder of script
# file_directory = folder_directory.joinpath("Liste til telefonnumre - Tilfredshedsundersøgelse.xlsx") # gets file and joins it with folder path
# excel_file = file_directory

excel_file = ''
first_row = 8
first_col = 1
url = 'https://study.epinionglobal.com/ta_e/kattegatcentret?abs=1&seg=1&test=1'



# ____________________________________ MAIN FUNCTIONS (CALLED FROM PYSIMPLEGUI)_________________________________________________

# def import_and_parse_numbers(excel_file, first_row, first_col):
#     phone_numbers = import_numbers_from_excel_as_list(excel_file, first_row, first_col)
#     parsed_phone_numbers = parse_numbers(phone_numbers) # Parses numbers in to list of tuples containing national number and country code.
#     print("SAMLET ANTAL TELEFONNUMRE: " + str(len(parsed_phone_numbers)) + "\nTELEFONNUMRE (telefonnummer, landekode): \n"+ str(parsed_phone_numbers))
#     return parsed_phone_numbers

    
# def send_surveys(url, parsed_phone_numbers):
#     driver = initialize_driver()
#     open_webform(driver, url)
#     delay = 5 # delay for click_buttonScripts
#     sent_to = []

#     for i in parsed_phone_numbers:
#         languageButtonSearchValue = get_lan_button_search_value(i)
#         click_button(driver, languageButtonSearchValue, delay)
#         fill_text_field(driver, '//input[@id="_Q1_O1"]', i[0]
#                       )
#         scroll_to_bottom(driver)
#         click_button(driver, '//input[@value="Send"]', delay)
#         time.sleep(1) # needed for reload of page
#         sent_to.append(i)
#     print_sent_to_report(sent_to)
#     driver.quit() # Quits the driver after loop is ended

def ClearSheet(excel_file, first_row, first_col):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    for col in range(first_col, sheet.max_column+1):
        for row in range(first_row, sheet.max_row+1):
            cell = sheet.cell(row, col)
            cell.value = None
    wb.save(excel_file)
    wb.close()
    print("Excel-ark er ryddet")


#_______________________________________ SUB FUNCTIONS (CALLED FROM  MAIN FUNCTIONS)__________________________________________

# def import_numbers_from_excel_as_list(excel_file, first_row, first_col):
#     # Sets workbook and sheet for openpyxl
#     wb = openpyxl.load_workbook(excel_file)
#     sheet = wb.active # gets active/first sheet-name of workbook
#     phone_numbers = [] # resets list
#     # loops through specified columns and rows.
#     for col in range(first_col, sheet.max_column+1):
#         for row in range(first_row, sheet.max_row+1):
#             cell = sheet.cell(row, col)
#             if cell.value: #and type(cell.value) == int:
#                 phone_numbers.append(str(cell.value)) # appends cell.value to phonenumbers - converts to string to have the plus-sign. (if saved as int, the plus-sign is removed, which makes it impossible for parse_numbers to determine if countrycode)
#     wb.close() # closes workbook
#     return phone_numbers

# def parse_numbers(phone_numbers):
#     # Parses phone_numbers into national number and country codesaved as tuples in parsed_numbers. Needs +-sign in front of country code. 
#     parsed_phone_numbers = []
#     phone_numbers_string = map(str, phone_numbers) # Makes sure that all numbers are converted to string to be used by phonenumbers.parse. (is converted in import_numbers_from_excel_as_list also)

#     # loops through phonenumbers and parses them
#     for i in phone_numbers_string:
#         if '+' in i: # If country code (signified by use of +(!)).
#             parsed_number = phonenumbers.parse(i, None) # No need for second argument, as country code is present in i.)
#             parsed_phone_numbers.append((parsed_number.national_number, parsed_number.country_code)) # appends national number and country code as tuple to parsed_phone_numbers
#         else: # If no country code is present signified by lack of "+"
#             parsed_phone_numbers.append(((i), 45)) # if no country code is present, only national number is appended to parsed_phone_numbers.
#     return parsed_phone_numbers


# def initialize_driver():
# # Path to the geckodriver executable
#     geckodriver_path = 'path/to/geckodriver'
#     # Find the Firefox binary location using the 'shutil' module
#     firefox_binary_path = shutil.which('firefox')
#     # Initialize the Firefox driver using a Service object and FirefoxOptions
#     service = Service(executable_path=geckodriver_path)
#     firefox_options = Options()
#     firefox_options.binary_location = firefox_binary_path
#     driver = webdriver.Firefox(service=service, options=firefox_options)
#     return driver

# def open_webform(driver, url):
#     driver.get(url) # Open the webform
#     #driver.maximize_window() # send-button is obscured by smaller window sizes


# def get_lan_button_search_value(i):
#     searchValue = ''
#     if i[1] == 45: # if country code is 45 (Denmark)
#         searchValue = '//input[@value="DAN"]'
#     elif i[1] == 49 or 43 or 41: # if 49 = Germany, 43 = Austria, 41 = Switzerland
#         searchValue = '//input[@value="DEU"]'
#     elif i[1] != 45 or 49 or 43 or 41 : # if country code is not 45 (Denmark)
#         searchValue = '//input[@value="ENG"]'
#     return searchValue

# # def ChooseLanguage():

# # def DetermineAreaCode():
    
# # def SelectAreaCodeFromDropdown():


# def click_button(driver, xpath, delay):
#     try:
#         element_present = expected_conditions.presence_of_element_located((By.XPATH, xpath))
#         WebDriverWait(driver, delay).until(element_present) # delays until element is present
#     except TimeoutException:
#        print("Timed out waiting for page to load") 
#     else:
#         button_to_click = driver.find_element(By.XPATH, xpath)
#         driver.execute_script("arguments[0].click();", button_to_click) # uses javascript to execute click. (Seleniums click function, needed to have button in view, which was problematic.)

# def fill_text_field(driver, xpath, phone_number):
#     # Find the text field for phone number and fill it with the phone number
#     text_field = driver.find_element(By.XPATH, xpath)
#     text_field.clear()
#     text_field.send_keys(phone_number)

# def scroll_to_bottom(driver):
#     driver.execute_script("window.scrollTo(0, document.body.scrollHeight);") # scroll to the bottom of screen

# def print_sent_to_report(sent_to):
#     print("---\nAntal sendte spørgeskemaer: " + str(len(sent_to)) + "\nSendt til følgende telefonnumre: " + str(sent_to))




#_________________________________________________________PYSIMPLEGUI_______________________________________________________


    # Create the PySimpleGUI layout
layout = [
    [sg.Text('Excel fil sti: '), sg.Input(excel_file, key= '-FILEPATH-'), sg.FileBrowse()],
    [sg.Text('Første række med telefonnumre: '), sg.InputText(first_row, key="-FIRSTROW-")],
    [sg.Text('Første kolonne med telefonnumre: '), sg.InputText(first_col, key="-FIRSTCOLUMN-")],
    [sg.Button('Importer', key="-IMPORT-")],
    [sg.Output(size=(80,20), key="-OUTPUT-")],
    [sg.Text("Vælg link:"), sg.Combo(["Test link", "Live link"], key="-COMBO-")],
    [sg.Button('Send', key="-SEND-"), sg.Button('Slet telefonnumre fra fil', key="-DELETE-"), sg.Button('Exit', key="-EXIT-")]
]

# Create the window with PySimpleGUI
window = sg.Window('Survey Sender', layout)

# Event loop to process events and update window
while True:
    event, values = window.read()
    # Ends program
    if event == sg.WIN_CLOSED or event == "-EXIT-":
        break
    # Calls import_and_parse_numbers when button Import is pressed
    elif event == "-IMPORT-":
        try:    
            excel_file = values['-FILEPATH-'] 
            first_row = int(values['-FIRSTROW-'])
            first_column = int(values['-FIRSTCOLUMN-'])
            parsed_phone_numbers = phonenumber_extractor.import_and_parse_numbers(excel_file, first_row, first_column)
        except PermissionError:
            print("Der er ikke adgang til excel-filen. Luk excel-filen og prøv igen.")  
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info() # defines exception-information
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(f"An error occurred: {e}{exc_type}{fname}{exc_tb.tb_lineno}")

    elif event == "-DELETE-":
        try:
            excel_file = values['-FILEPATH-'] 
            first_row = int(values['-FIRSTROW-'])
            first_column = int(values['-FIRSTCOLUMN-'])
            # Display a confirmation prompt before executing the script
            confirm = sg.popup_yes_no("Er du sikker på, at du vil slette telefonnumrene fra excel-arket?", title="Bekræftelse")
            if confirm == "Yes":
                ClearSheet(excel_file, first_row, first_column)
                sg.popup("Cellerne er blevet slettet.", title="Færdig")
        except PermissionError:
            print("Der er ikke adgang til excel-filen. Luk excel-filen og prøv igen.")          
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info() # defines exception-information
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(f"An error occurred: {e}{exc_type}{fname}{exc_tb.tb_lineno}")
        

    # Calls send_surveys-function with updated variables
    elif event == "-SEND-":
        try:
            # checks for selection of combo. Sets url to appropriate choice.
            if values["-COMBO-"] == '':
                raise TypeError("Intet link valgt")
            elif values["-COMBO-"] == "Test link": 
                url = "https://study.epinionglobal.com/ta_e/kattegatcentret?abs=1&seg=1&test=1"
            elif values["-COMBO-"] == "Live link":
                url = "https://study.epinionglobal.com/ta_e/kattegatcentret?abs=1&seg=1"
            webform_filler.send_surveys(url, parsed_phone_numbers)
        except PermissionError:
            print("Der er ikke adgang til excel-filen. Luk excel-filen og prøv igen.")    
        except TypeError:
            print("Ingen data importeret. Vælg en gyldig filsti og prøv igen.")
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info() # defines exception-information
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(f"Der er opstået en fejl. Send et billede af følgende fejlmeddelelse til Victor: {e}{exc_type}{fname}{exc_tb.tb_lineno}")