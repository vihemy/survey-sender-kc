# External libraries
import openpyxl
import re
from phonenumbers import NumberParseException

# Internal libraries
from phonenumber_class import PhoneNumber
import printer


def check_if_cell_content_is_valid(cell):
    """Check if cell content is valid. Returns True if valid, False if not."""
    if check_if_cell_is_empty(cell):
        return False
    elif check_if_cell_contains_less_than_8_digits(cell):
        return False
    elif check_if_cell_contains_inappropriate_characters(cell):
        return False
    else:
        return True


def check_if_cell_is_empty(cell):
    """Check if cell is empty. Returns True if empty, False if not."""
    return bool(cell.value == None)


def check_if_cell_contains_less_than_8_digits(cell):
    """Check if cell contains less than 8 digits. Returns True if less than 8 digits, False if not."""
    return bool(len(cell.value) < 8)


def check_if_cell_contains_inappropriate_characters(cell):
    """Check if cell contains characters other than numbers and "+". Returns True if inappropriate characters, False if not."""
    # checks if cell value contains anything but numbers and '+'
    return bool(re.search(r'[^0-9+]', cell.value))


def import_numbers_from_excel_as_list(excel_file, first_row, first_col):
    """Import phone numbers from excel file and return as list of PhoneNumber-objects"""
    phone_numbers = []  # resets list
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active  # gets active/first sheet-name of workbook

    # loops through specified columns and rows.
    for col in range(first_col, sheet.max_column+1):
        for row in range(first_row, sheet.max_row+1):
            cell = sheet.cell(row, col)
            # if cell.value is not empty and does not contains letters, add as phoenumber-object
            if check_if_cell_content_is_valid(cell):
                try:
                    # creates PhoneNumber-object
                    phone_number = PhoneNumber(cell.value)
                    # appends phone_number to list
                    phone_numbers.append(phone_number)
                except NumberParseException:
                    print(
                        f"Række {row}, kolonne {col} indeholder ugyldigt telefonnummer. Nummeret er ikke blevet importeret.")
                    continue

    wb.close()  # closes workbook
    printer.print_imported_numbers(phone_numbers)
    return phone_numbers
